from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import Count
from django.http import HttpResponse
from django.utils import formats, translation
from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as PDFImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from apps.dashboard.services import category_costs, expense_base_queryset, money_sum, monthly_costs, overview_totals
from apps.finance.models import ApprovalStatus
from apps.projects.models import Project
from apps.projects.selectors import filter_projects_for_user


ZERO = Decimal("0.00")
BRAND_NAME = "RAK"
APP_NAME = "Site Engineer Cost Management System"
BRAND_PRIMARY = "#a30f17"
BRAND_STEEL = "#64748b"
BRAND_LOGO_PATH = settings.BASE_DIR / "static" / "branding" / "rak-logo.png"

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
except ImportError:  # Arabic PDF shaping is enhanced when optional packages are installed.
    arabic_reshaper = None
    get_display = None


REPORT_TITLES = {
    "project-summary": "Project Summary",
    "monthly-cost": "Monthly Cost Report",
    "category-cost": "Category Cost Report",
    "vendor-statement": "Vendor Statement",
    "pending-payments": "Pending Payments Report",
    "cash-flow": "Cash Flow Report",
}

HEADER_LABELS = {
    "amount": "Amount",
    "balance": "Balance",
    "cash_balance": "Cash Balance",
    "cash_in": "Cash In",
    "cash_out": "Cash Out",
    "category__name": "Category",
    "category_breakdown": "Category Breakdown",
    "client_name": "Client",
    "code": "Code",
    "days_pending": "Days Pending",
    "expense": "Expense",
    "expense_count": "Expense Count",
    "expense_date": "Expense Date",
    "expense_id": "Expense ID",
    "expenses": "Expenses",
    "month": "Month",
    "paid": "Paid",
    "payment_status": "Payment Status",
    "pending": "Pending",
    "percentage_of_total_cost": "Percentage of Total Cost",
    "project": "Project",
    "project__name": "Project",
    "project_cost_gap": "Project Cost Gap",
    "project_id": "Project ID",
    "status": "Status",
    "total": "Total",
    "total_cash_in": "Total Cash In",
    "total_cash_out": "Total Cash Out",
    "total_paid": "Total Paid",
    "total_pending": "Total Pending",
    "vendor": "Vendor",
    "vendor__name": "Vendor",
    "vendor_id": "Vendor ID",
}

STATUS_LABELS = {
    "ACTIVE": "Active",
    "APPROVED": "Approved",
    "CANCELLED": "Cancelled",
    "COMPLETED": "Completed",
    "DRAFT": "Draft",
    "ON_HOLD": "On hold",
    "PAID": "Paid",
    "PARTIALLY_PAID": "Partially paid",
    "REJECTED": "Rejected",
    "SUBMITTED": "Submitted",
    "UNPAID": "Unpaid",
}

MONEY_FIELDS = {
    "amount",
    "balance",
    "cash_balance",
    "cash_in",
    "cash_out",
    "paid",
    "pending",
    "project_cost_gap",
    "total",
    "total_cash_in",
    "total_cash_out",
    "total_paid",
    "total_pending",
}

ARABIC_FONT_NAME = "ArabicReportFont"
ARABIC_FONT_PATHS = [
    Path("C:/Windows/Fonts/arial.ttf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf"),
    Path("/usr/share/fonts/truetype/freefont/FreeSans.ttf"),
]


def report_params(request):
    return request.query_params


def active_language():
    return (translation.get_language() or "en").split("-")[0]


def is_rtl():
    return active_language() == "ar"


def report_title(report_name):
    return _(REPORT_TITLES.get(report_name, report_name.replace("-", " ").title()))


def header_label(header):
    return _(HEADER_LABELS.get(header, header.replace("_", " ").title()))


def status_label(value):
    return _(STATUS_LABELS.get(str(value), str(value).replace("_", " ").title()))


def localized_number(value, decimal_pos=2):
    return formats.number_format(value, decimal_pos=decimal_pos, use_l10n=True, force_grouping=True)


def localized_month(value):
    try:
        parsed = datetime.strptime(str(value), "%Y-%m").date()
    except ValueError:
        return str(value)
    return formats.date_format(parsed, "YEAR_MONTH_FORMAT", use_l10n=True)


def format_export_value(key, value):
    if value is None:
        return ""
    if key == "month":
        return localized_month(value)
    if key in {"status", "approval_status", "payment_status"}:
        return status_label(value)
    if key == "percentage_of_total_cost":
        return f"{localized_number(value)}%"
    if key.endswith("_date") and isinstance(value, (date, datetime)):
        return formats.date_format(value, "DATE_FORMAT", use_l10n=True)
    if key.endswith("_date") and isinstance(value, str):
        try:
            return formats.date_format(datetime.fromisoformat(value).date(), "DATE_FORMAT", use_l10n=True)
        except ValueError:
            return value
    if key in MONEY_FIELDS:
        return localized_number(value)
    if isinstance(value, Decimal):
        return localized_number(value)
    if isinstance(value, int):
        return formats.number_format(value, decimal_pos=0, use_l10n=True, force_grouping=True)
    return value


def register_report_font():
    if not is_rtl():
        return "Helvetica", "Helvetica-Bold"
    if ARABIC_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
        for font_path in ARABIC_FONT_PATHS:
            if font_path.exists():
                pdfmetrics.registerFont(TTFont(ARABIC_FONT_NAME, str(font_path)))
                break
    if ARABIC_FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return ARABIC_FONT_NAME, ARABIC_FONT_NAME
    return "Helvetica", "Helvetica-Bold"


def shape_pdf_text(value):
    text = str(value)
    if not is_rtl() or not any("\u0600" <= char <= "\u06ff" for char in text):
        return text
    if arabic_reshaper and get_display:
        return get_display(arabic_reshaper.reshape(text))
    return text


def brand_logo_path():
    return BRAND_LOGO_PATH if BRAND_LOGO_PATH.exists() else None


def project_summary_report(user, params):
    projects = filter_projects_for_user(Project.objects.all(), user)
    project_id = params.get("project")
    if project_id:
        projects = projects.filter(id=project_id)
    rows = []
    for project in projects:
        totals = overview_totals(user, {"project": project.id, **params})
        rows.append(
            {
                "project_id": project.id,
                "project": project.name,
                "code": project.code,
                "client_name": project.client_name,
                "status": project.status,
                **totals,
                "category_breakdown": category_costs(user, {"project": project.id, **params}),
            }
        )
    return rows


def monthly_cost_report(user, params):
    return monthly_costs(user, params)


def category_cost_report(user, params):
    rows = category_costs(user, params)
    total = sum((row["total"] for row in rows), ZERO)
    for row in rows:
        row["pending"] = row["total"] - row["paid"]
        row["percentage_of_total_cost"] = (row["total"] / total * 100) if total else ZERO
    return rows


def vendor_statement_report(user, params):
    expenses = expense_base_queryset(user, params).filter(vendor__isnull=False)
    vendor_id = params.get("vendor")
    if vendor_id:
        expenses = expenses.filter(vendor_id=vendor_id)
    vendors = (
        expenses.values("vendor_id", "vendor__name")
        .annotate(total=money_sum("total_amount"), paid=money_sum("paid_amount"), pending=money_sum("pending_amount"), expense_count=Count("id"))
        .order_by("vendor__name")
    )
    rows = []
    for vendor in vendors:
        related = expenses.filter(vendor_id=vendor["vendor_id"]).values(
            "id",
            "project__name",
            "description",
            "expense_date",
            "total_amount",
            "paid_amount",
            "pending_amount",
            "approval_status",
            "payment_status",
        )
        rows.append({**vendor, "expenses": list(related)})
    return rows


def pending_payments_report(user, params):
    today = timezone.localdate()
    expenses = expense_base_queryset(user, params).filter(pending_amount__gt=0, approval_status=ApprovalStatus.APPROVED)
    rows = []
    for expense in expenses.select_related("project", "vendor"):
        rows.append(
            {
                "expense_id": expense.id,
                "expense": expense.description,
                "vendor": expense.vendor.name if expense.vendor else "",
                "project": expense.project.name,
                "total": expense.total_amount,
                "paid": expense.paid_amount,
                "pending": expense.pending_amount,
                "days_pending": (today - expense.expense_date).days,
            }
        )
    return rows


def cash_flow_report(user, params):
    rows = defaultdict(lambda: {"month": "", "cash_in": ZERO, "cash_out": ZERO, "paid": ZERO, "pending": ZERO, "balance": ZERO})
    for row in monthly_costs(user, params):
        month = row["month"]
        rows[month]["month"] = month
        rows[month]["cash_in"] = row["cash_in"]
        rows[month]["cash_out"] = row["cash_out"]
        rows[month]["paid"] = row["paid"]
        rows[month]["pending"] = row["pending"]
        rows[month]["balance"] = row["cash_in"] - row["paid"]
    return [rows[key] for key in sorted(rows.keys())]


REPORT_BUILDERS = {
    "project-summary": project_summary_report,
    "monthly-cost": monthly_cost_report,
    "category-cost": category_cost_report,
    "vendor-statement": vendor_statement_report,
    "pending-payments": pending_payments_report,
    "cash-flow": cash_flow_report,
}


def flatten_row(row):
    flattened = {}
    for key, value in row.items():
        if isinstance(value, list):
            flattened[key] = len(value)
        else:
            flattened[key] = value
    return flattened


def build_excel_response(user, params):
    report_name = params.get("report", "project-summary")
    data = REPORT_BUILDERS.get(report_name, project_summary_report)(user, params)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_title(report_name)[:31]
    sheet.sheet_view.rightToLeft = is_rtl()
    logo_path = brand_logo_path()
    title_column = "B" if logo_path else "A"
    if logo_path:
        try:
            logo = ExcelImage(str(logo_path))
            logo.width = 72
            logo.height = 50
            sheet.add_image(logo, "A1")
        except Exception:
            title_column = "A"
    sheet[f"{title_column}1"] = BRAND_NAME
    sheet[f"{title_column}2"] = _(APP_NAME)
    sheet[f"{title_column}3"] = report_title(report_name)
    sheet[f"{title_column}1"].font = Font(bold=True, color=BRAND_PRIMARY.replace("#", ""), size=14)
    sheet[f"{title_column}2"].font = Font(color=BRAND_STEEL.replace("#", ""))
    sheet[f"{title_column}3"].font = Font(bold=True, color="111827")
    sheet.append([])
    rows = [flatten_row(row) for row in data]
    if rows:
        headers = list(rows[0].keys())
        header_row_index = sheet.max_row + 1
        sheet.append([header_label(header) for header in headers])
        for row in rows:
            sheet.append([format_export_value(header, row.get(header, "")) for header in headers])
    else:
        header_row_index = sheet.max_row + 1
        sheet.append([_("No data")])

    alignment = Alignment(horizontal="right" if is_rtl() else "left")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment
    for cell in sheet[header_row_index]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND_PRIMARY.replace("#", ""))
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max([len(value) for value in values] + [12]) + 2, 34)
        sheet.column_dimensions[column_cells[0].column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{report_name}.xlsx"'
    return response


def build_pdf_response(user, params):
    report_name = params.get("report", "project-summary")
    data = REPORT_BUILDERS.get(report_name, project_summary_report)(user, params)
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=28, leftMargin=28)
    styles = getSampleStyleSheet()
    font_name, bold_font_name = register_report_font()
    title_style = ParagraphStyle(
        "LocalizedTitle",
        parent=styles["Title"],
        alignment=2 if is_rtl() else 1,
        fontName=font_name,
        textColor=colors.HexColor("#111827"),
    )
    brand_style = ParagraphStyle(
        "LocalizedBrand",
        parent=styles["BodyText"],
        alignment=2 if is_rtl() else 0,
        fontName=bold_font_name,
        fontSize=10,
        textColor=colors.HexColor(BRAND_PRIMARY),
    )
    subtitle_style = ParagraphStyle(
        "LocalizedSubtitle",
        parent=styles["BodyText"],
        alignment=2 if is_rtl() else 0,
        fontName=font_name,
        fontSize=8,
        textColor=colors.HexColor(BRAND_STEEL),
    )
    body_style = ParagraphStyle(
        "LocalizedBody",
        parent=styles["BodyText"],
        alignment=2 if is_rtl() else 0,
        fontName=font_name,
        fontSize=8,
    )
    story = []
    logo_path = brand_logo_path()
    if logo_path:
        story.append(PDFImage(str(logo_path), width=72, height=50, hAlign="RIGHT" if is_rtl() else "LEFT"))
        story.append(Spacer(1, 4))
    story.extend(
        [
            Paragraph(shape_pdf_text(BRAND_NAME), brand_style),
            Paragraph(shape_pdf_text(_(APP_NAME)), subtitle_style),
            Spacer(1, 10),
            Paragraph(shape_pdf_text(report_title(report_name)), title_style),
            Spacer(1, 12),
        ]
    )
    rows = [flatten_row(row) for row in data]
    if rows:
        headers = list(rows[0].keys())
        display_headers = list(reversed(headers)) if is_rtl() else headers
        table_data = [[Paragraph(shape_pdf_text(header_label(header)), body_style) for header in display_headers]]
        for row in rows[:100]:
            table_data.append([Paragraph(shape_pdf_text(format_export_value(header, row.get(header, ""))), body_style) for header in display_headers])
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_PRIMARY)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
                    ("FONTNAME", (0, 1), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (-1, -1), "RIGHT" if is_rtl() else "LEFT"),
                ]
            )
        )
        story.append(table)
    else:
        story.append(Paragraph(shape_pdf_text(_("No data")), body_style))
    def draw_footer(canvas, document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor(BRAND_PRIMARY))
        canvas.setLineWidth(0.5)
        canvas.line(document.leftMargin, 24, A4[0] - document.rightMargin, 24)
        canvas.setFillColor(colors.HexColor(BRAND_STEEL))
        canvas.setFont(font_name, 7)
        footer = shape_pdf_text(f"{BRAND_NAME} | {_(APP_NAME)}")
        page = shape_pdf_text(_("Page %(page)s") % {"page": document.page})
        if is_rtl():
            canvas.drawRightString(A4[0] - document.rightMargin, 14, footer)
            canvas.drawString(document.leftMargin, 14, page)
        else:
            canvas.drawString(document.leftMargin, 14, footer)
            canvas.drawRightString(A4[0] - document.rightMargin, 14, page)
        canvas.restoreState()

    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{report_name}.pdf"'
    return response
